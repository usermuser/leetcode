from openpyxl import load_workbook

# Открываем файл Excel
workbook = load_workbook(filename='loans.xlsx')

# Выбираем активный лист (можно указать имя листа вместо 'Sheet1')
sheet = workbook.active

# Проходим по всем строкам в столбце C и удаляем первый символ
total = 0
for row in sheet.iter_rows(min_row=2, min_col=3, max_col=5, values_only=True):
    if row[0] is not None and len(row[0]) > 1:
        # row[0] = row[0][1:]
        # new_row = row[0][1:]
        # print(new_row)

        date = row[1]
        if date.year == 2021 and int(row[0][-1]) == 0:
            total += 1

# Сохраняем изменения в файл
workbook.save(filename='loans.xlsx')
print(total)
