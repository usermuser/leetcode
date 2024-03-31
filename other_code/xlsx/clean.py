from openpyxl import load_workbook

# Открываем файл Excel
workbook = load_workbook(filename='loans.xlsx')

# Выбираем активный лист (можно указать имя листа вместо 'Sheet1')
loans = workbook["Loans"]

# Проходим по всем строкам в столбце C и удаляем первый символ
total_amount = 0
rating_threshold = 12
companies = {}  # bid: amount
for row in loans.iter_rows(min_row=2, min_col=2, max_col=5, values_only=True):

    print(row[0], row[1], row[2])
    bid = row[0]
    rating = row[1]
    amount = row[2]

    if rating <= rating_threshold:
        total_amount += amount
        # if bid not in companies:
        #     companies[bid] = amount
        # else:
        #     companies[bid] = amount + companies[bid]


print(total_amount)
